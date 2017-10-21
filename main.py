#!/bin/python
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import os

#sorts through data in given format from files in current directory with the
#defined file extension, searches for defined element isotopes
DATA_FILE_EXTENSION = ".plt"
ELEMENT = "pu"
OUTPUT_FILE = ELEMENT + " raw data.xlsx"

#sort through a list of files, return files with the given extension
def sort_file_extensions(files, extension):
    return [i for i in files if i[-len(extension):] == extension]

def get_file_lines(filename):
    with open(filename, "r") as open_file: #open file to read and close when donw
        return open_file.readlines()

#get number from "Depletion material no.    x" for row1
def row1_data(file_lines):
    return file_lines[0].split()[3][:-1]

def col_b_data(file_lines):
    return ["total mass of unit", "burnup(days)"] + file_lines[5].replace("e", "E").split() #e replacement because notation

#get element data for columns
def get_col_element_data(file_lines, element):
    output = []
    for line in file_lines[6:]:
        if line.split()[0][:len(element)] == element:
            output = output + [line.strip().replace("e", "E").split()] #e replacement because notation
    output[0] = [row1_data(file_lines)] + output[0]

    return output

def main():
    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = ELEMENT + " raw data"

    directory_files = os.listdir("./") #get files in current directory

    current_column = 0

    for i, data_file in enumerate(sort_file_extensions(directory_files, DATA_FILE_EXTENSION)):
        #set column b
        file_lines = get_file_lines(data_file)
        if i == 0:
            col_b = col_b_data(file_lines)
            for row in range(len(col_b)):
                sheet1.cell(column=2, row=1+row, value=col_b[row])

        #set other columns
        col_element_data = get_col_element_data(file_lines, ELEMENT)
        for col in range(len(col_element_data)):
            for row in range(len(col_element_data[col])):
                if (col == 0):
                    sheet1.cell(column=3+col+current_column, row=1+row, value=col_element_data[col][row])
                else:
                    sheet1.cell(column=3+col+current_column, row=2+row, value=col_element_data[col][row])
        current_column += len(col_element_data)
    wb.save(filename=OUTPUT_FILE)
    
#execute main function if run as a program
if __name__ == "__main__":
    main()