#Written by Max Weis (maxrobweis@gmail.com)

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import os

#TODO
# add total

#sorts through data in given format from files in current directory with the
#defined file extension, searches for defined element isotopes
DATA_FILE_EXTENSION = ".plt"
ELEMENT = "pu"
OUTPUT_FILE = ELEMENT + " raw data.xlsx"

#sort through a list of files, return files with the given extension
def sort_file_extensions(files, extension):
    return [i for i in files if i[-len(extension):] == extension]

def get_file_lines(filename):
    with open(filename, "r") as open_file: #open file to read and close when donw
        return open_file.readlines()

#get number from "Depletion material no.    x" for row1
def row1_data(file_lines):
    return file_lines[0].split()[3][:-1]

def col_b_data(file_lines):
    return ["total mass of unit", "burnup(days)"] + file_lines[5].split()

#get element data for columns
def get_col_element_data(file_lines, element):
    output = []
    for line in file_lines[6:]:
        if line.split()[0][:len(element)] == element:
            output = output + [line.strip().split()]
    output[0] = [row1_data(file_lines)] + output[0]

    return output

def main():
    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = ELEMENT + " raw data"

    directory_files = os.listdir("./")
    directory_files = sort_file_extensions(directory_files, DATA_FILE_EXTENSION) #get files with desired extension
    directory_files.sort(key=lambda x: int(x.split(".")[1])) #sort files by numerical order for listing in row 1

    current_column = 0

    for i, data_file in enumerate(directory_files):
        #set column b
        file_lines = get_file_lines(data_file)
        if i == 0:
            col_b = col_b_data(file_lines)

            for row in range(len(col_b)):
                try:
                    value = float(col_b[row])
                except ValueError:
                    value = col_b[row]

                sheet1.cell(column=2, row=1+row, value=value)

        #set other columns
        col_element_data = get_col_element_data(file_lines, ELEMENT)
        for col in range(len(col_element_data)):
            for row in range(len(col_element_data[col])):
                try:
                    value = float(col_element_data[col][row])
                except ValueError:
                    value = col_element_data[col][row]

                if col == 0:
                    sheet1.cell(column=3+col+current_column, row=1+row, value=value)
                else:
                    sheet1.cell(column=3+col+current_column, row=2+row, value=value)
        current_column += len(col_element_data)
    wb.save(filename=OUTPUT_FILE)
    
if __name__ == "__main__":
    main()
